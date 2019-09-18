import openpyxl
import xlrd


class ExcelUtil():
    def noOfRows(fileName, sheet_name):
        wb = xlrd.open_workbook(fileName)
        sheet = wb.sheet_by_name(sheet_name)
        sheet.cell_value(0, 0)
        return sheet.nrows


    def noofcols(fileName, sheet_name):
        wb = xlrd.open_workbook(fileName)
        sheet = wb.sheet_by_name(sheet_name)
        sheet.cell_value(0, 0)
        return sheet.ncols

    def getcelldata(fileName, sheet_name, rownum, colno):
        wb = xlrd.open_workbook(fileName)
        sheet = wb.sheet_by_name(sheet_name)
        sheet.cell_value(0, 0)
        return sheet.cell_value(rownum, colno)


    def getdata(fileName, sheet_name):
        wb = xlrd.open_workbook(fileName)
        sheet = wb.sheet_by_name(sheet_name)
        sheet.cell_value(0, 0)
        rows = []
        for row_id in range (1, sheet.nrows):
            rows.append(sheet.row_values(row_id, 0, sheet.ncols))
        return rows


    def writedata(fileName, sheet_name, Row_num, Result):
        wb = openpyxl.load_workbook(fileName)
        sheet = wb[sheet_name]
        sheet["D"+str(int(Row_num)+1)] = Result
        wb.save(fileName)
        wb.close()

    def writedatatoMultipleCells(fileName, sheet_name,col_no, Row_num_start,Row_num_start_end, Result):
        wb = openpyxl.load_workbook(fileName)
        sheet = wb[sheet_name]
        for i in range(Row_num_start, Row_num_start_end):
            sheet[col_no+str(int(i)+1)] = Result
        wb.save(fileName)
        wb.close()

    def writedataSingleCell(fileName, sheet_name,col_no, Row_num, Result):
        wb = openpyxl.load_workbook(fileName)
        sheet = wb[sheet_name]
        sheet[col_no+str(int(Row_num)+1)] = Result
        wb.save(fileName)
        wb.close()

    def addSheet(fileName, sheet_name):
        wb = openpyxl.load_workbook(fileName)
        wb.create_sheet(sheet_name)
        wb.save(fileName)
        wb.close()






